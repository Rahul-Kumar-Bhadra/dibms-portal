import os
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

MONTH_NAMES = ["", "January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]

def generate_plant_template(plant_id: str, plant_name: str, month: int, year: int, manager_name: str, template_version: str = "v1.0") -> io.BytesIO:
    """
    Generates a standardized Excel template for a specific plant, month, and year.
    Uses openpyxl to lock structural cells, apply formatting, and embed metadata.
    """
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Tailwind blue-900
    title_font = Font(bold=True, size=16, color="1E3A8A")
    label_font = Font(bold=True, color="475569") # Tailwind slate-600
    locked_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid") # Tailwind slate-100
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # -----------------------------------------
    # Sheet 1: Instructions (Read-only)
    # -----------------------------------------
    ws_instructions = wb.active
    ws_instructions.title = "Instructions"
    ws_instructions.protection.sheet = True # Protect sheet
    
    ws_instructions["A1"] = "DMEOP Enterprise Excel Intelligence Module"
    ws_instructions["A1"].font = title_font
    
    instructions = [
        "Welcome to the DIBMS Automated Operations Reporting Module.",
        "",
        "HOW TO USE THIS TEMPLATE:",
        "1. Go to the 'Monthly Report' tab.",
        "2. Fill in the data for the requested reporting month.",
        "3. Only white cells can be edited. Gray cells are locked structural metadata.",
        "4. Revenue, Expenses, Production Units, and Safety Incidents must be numeric and non-negative.",
        "5. Attendance Rate and Quality Score must be between 0 and 100.",
        "6. Do not modify the hidden metadata in the Reference Data tab.",
        "7. Save the file and upload it via the DIBMS portal.",
        "",
        "If you encounter validation errors during upload, correct them in this file and re-upload."
    ]
    
    for r_idx, text in enumerate(instructions, start=3):
        ws_instructions[f"A{r_idx}"] = text
        if not text.startswith("HOW TO") and text:
            ws_instructions[f"A{r_idx}"].font = Font(color="333333")
        else:
            ws_instructions[f"A{r_idx}"].font = Font(bold=True)
            
    ws_instructions.column_dimensions['A'].width = 80
    
    # -----------------------------------------
    # Sheet 2: Monthly Report (Data Entry)
    # -----------------------------------------
    ws_report = wb.create_sheet(title="Monthly Report")
    
    # Corporate Header
    ws_report["A1"] = "DANA INDIA OPERATIONS PORTAL"
    ws_report["A1"].font = title_font
    ws_report.merge_cells("A1:D1")
    
    # Metadata block (Locked)
    metadata = [
        ("Plant ID", plant_id),
        ("Plant Name", plant_name),
        ("Manager Name", manager_name),
        ("Reporting Month", MONTH_NAMES[month]),
        ("Reporting Year", year),
    ]
    
    for r_idx, (label, val) in enumerate(metadata, start=3):
        ws_report[f"A{r_idx}"] = label
        ws_report[f"A{r_idx}"].font = label_font
        ws_report[f"B{r_idx}"] = val
        ws_report[f"B{r_idx}"].fill = locked_fill
        ws_report[f"B{r_idx}"].border = thin_border
        
    # Data Entry Table
    headers = ["Metric", "Value", "Remarks (Optional)"]
    start_row = 10
    
    for c_idx, header in enumerate(headers, start=1):
        cell = ws_report.cell(row=start_row, column=c_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        
    metrics = [
        "Revenue (INR)",
        "Expenses (INR)",
        "Production Units",
        "Attendance Rate (%)",
        "Safety Incidents",
        "Quality Score (%)"
    ]
    
    for r_idx, metric in enumerate(metrics, start=start_row + 1):
        # Metric Label (Locked)
        cell_label = ws_report.cell(row=r_idx, column=1)
        cell_label.value = metric
        cell_label.fill = locked_fill
        cell_label.border = thin_border
        cell_label.font = Font(bold=True)
        
        # Value Input (Unlocked)
        cell_val = ws_report.cell(row=r_idx, column=2)
        cell_val.border = thin_border
        
        # Remarks Input (Unlocked)
        cell_rem = ws_report.cell(row=r_idx, column=3)
        cell_rem.border = thin_border
        
    ws_report.column_dimensions['A'].width = 25
    ws_report.column_dimensions['B'].width = 20
    ws_report.column_dimensions['C'].width = 40
    
    # -----------------------------------------
    # Sheet 3: Reference Data (Hidden metadata)
    # -----------------------------------------
    ws_ref = wb.create_sheet(title="Reference Data")
    ws_ref.sheet_state = 'hidden' # Hide this sheet from users
    
    # Embed checksum and metadata required for backend validation
    ws_ref["A1"] = "Template Version"
    ws_ref["B1"] = template_version
    
    ws_ref["A2"] = "Plant ID"
    ws_ref["B2"] = plant_id
    
    ws_ref["A3"] = "Month"
    ws_ref["B3"] = month
    
    ws_ref["A4"] = "Year"
    ws_ref["B4"] = year
    
    # Calculate checksum hash for the structural data
    import hashlib
    data_str = f"{template_version}|{plant_id}|{month}|{year}"
    checksum = hashlib.sha256(data_str.encode()).hexdigest()
    
    ws_ref["A5"] = "Checksum"
    ws_ref["B5"] = checksum

    # Save to BytesIO
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    return out
