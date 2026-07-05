import os
from typing import Any, List, Optional
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session

from app.api import deps
from app.models import PlantDocument, User, AuditLog, Plant
from app.schemas import DocumentOut
from app.core.config import settings

router = APIRouter()

@router.get("/", response_model=List[DocumentOut])
def read_documents(
    plant_id: Optional[str] = None,
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_user)
) -> Any:
    query = db.query(PlantDocument)
    
    # RBAC check
    if current_user.role == "Plant Manager":
        query = query.filter(PlantDocument.plant_id == current_user.plant_id)
    elif plant_id:
        query = query.filter(PlantDocument.plant_id == plant_id)
        
    docs = query.all()
    
    for d in docs:
        d.plant_name = d.plant.name
        if d.uploaded_by:
            uploader = db.query(User).filter(User.id == d.uploaded_by).first()
            if uploader:
                d.uploaded_by_name = f"{uploader.first_name or ''} {uploader.last_name or ''}".strip() or uploader.email
            else:
                d.uploaded_by_name = "Unknown"
        else:
            d.uploaded_by_name = "System"
            
    return docs

@router.post("/upload", response_model=DocumentOut)
def upload_document(
    plant_id: str,
    category: str = Form("SOP"),
    file: UploadFile = File(...),
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_user)
) -> Any:
    # RBAC check
    if current_user.role == "Plant Manager" and plant_id != current_user.plant_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You can only upload files for your own manufacturing plant."
        )
        
    plant = db.query(Plant).filter(Plant.id == plant_id).first()
    if not plant:
        raise HTTPException(status_code=404, detail="Plant not found")
        
    plant_upload_dir = os.path.join(settings.UPLOAD_DIR, plant_id)
    os.makedirs(plant_upload_dir, exist_ok=True)
    
    file_path = os.path.join(plant_upload_dir, file.filename)
    
    contents = file.file.read()
    file_size = len(contents)
    
    with open(file_path, "wb") as f:
        f.write(contents)
        
    ext = os.path.splitext(file.filename)[1].lower()
    if ext in [".xlsx", ".xls", ".csv"]:
        file_type = "Excel"
    elif ext in [".pdf"]:
        file_type = "PDF"
    elif ext in [".png", ".jpg", ".jpeg", ".webp"]:
        file_type = "Image"
    else:
        file_type = "Document"
        
    db_doc = PlantDocument(
        plant_id=plant_id,
        file_name=file.filename,
        file_path=f"uploads/{plant_id}/{file.filename}",
        file_type=file_type,
        file_size=file_size,
        category=category,
        uploaded_by=current_user.id
    )
    db.add(db_doc)
    
    db.add(AuditLog(
        user_id=current_user.id,
        action="Upload Technical Document",
        details=f"Uploaded file {file.filename} ({file_type}) under category {category} for plant {plant_id}."
    ))
    db.commit()
    db.refresh(db_doc)
    
    db_doc.plant_name = plant.name
    db_doc.uploaded_by_name = f"{current_user.first_name or ''} {current_user.last_name or ''}".strip() or current_user.email
    return db_doc

@router.get("/download/{doc_id}")
def download_document(
    doc_id: str,
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_user)
) -> Any:
    doc = db.query(PlantDocument).filter(PlantDocument.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
        
    if current_user.role == "Plant Manager" and doc.plant_id != current_user.plant_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You cannot download files of other manufacturing plants."
        )
        
    full_path = doc.file_path.replace("/", os.sep)
    
    if not os.path.exists(full_path):
        os.makedirs(os.path.dirname(full_path), exist_ok=True)
        if doc.file_name.endswith('.xlsx'):
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Document"
            ws["A1"] = "Simulated Placeholder File"
            ws["A2"] = f"File Name: {doc.file_name}"
            ws["A3"] = f"Plant ID: {doc.plant_id}"
            wb.save(full_path)
        else:
            with open(full_path, "w") as f:
                f.write(f"This is a simulated placeholder file for {doc.file_name} in plant {doc.plant_id}.")
                
    db.add(AuditLog(
        user_id=current_user.id,
        action="Download Technical Document",
        details=f"Downloaded file {doc.file_name} from plant {doc.plant_id}."
    ))
    db.commit()
    
    media_type = "application/octet-stream"
    if doc.file_name.lower().endswith(".xlsx"):
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif doc.file_name.lower().endswith(".pdf"):
        media_type = "application/pdf"
        
    return FileResponse(
        path=full_path,
        filename=doc.file_name,
        media_type=media_type
    )

@router.delete("/{doc_id}")
def delete_document(
    doc_id: str,
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_user)
) -> Any:
    doc = db.query(PlantDocument).filter(PlantDocument.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
        
    if current_user.role == "Plant Manager" and doc.plant_id != current_user.plant_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You cannot delete files of other manufacturing plants."
        )
        
    full_path = doc.file_path.replace("/", os.sep)
    if os.path.exists(full_path):
        try:
            os.remove(full_path)
        except Exception:
            pass
            
    db.delete(doc)
    db.add(AuditLog(
        user_id=current_user.id,
        action="Delete Technical Document",
        details=f"Deleted file {doc.file_name} from plant {doc.plant_id}."
    ))
    db.commit()
    return {"message": "Document deleted successfully"}
