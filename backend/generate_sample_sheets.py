import os
import io
import hashlib
import random
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.protection import SheetProtection

MONTH_NAMES = ["", "January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]

plants = [
    {"id": "PL001", "name": "Pune Plant", "manager": "Pune Plant Manager"},
    {"id": "PL002", "name": "Chennai Plant", "manager": "Chennai Plant Manager"},
    {"id": "PL003", "name": "Bengaluru Plant", "manager": "Bengaluru Plant Manager"},
    {"id": "PL004", "name": "Gurugram Plant", "manager": "Gurugram Plant Manager"},
    {"id": "PL005", "name": "Kolkata Plant", "manager": "Kolkata Plant Manager"},
    {"id": "PL006", "name": "Jamshedpur Plant", "manager": "Jamshedpur Plant Manager"},
    {"id": "PL007", "name": "Hyderabad Plant", "manager": "Hyderabad Plant Manager"},
    {"id": "PL008", "name": "Ahmedabad Plant", "manager": "Ahmedabad Plant Manager"},
]

month = 6
year = 2026
template_version = "v1.0"

remarks_pool = [
    "Production target achieved.",
    "Scheduled maintenance completed.",
    "Minor quality improvement required.",
    "Excellent safety performance.",
    "Strong production output.",
    "Preventive maintenance completed.",
    "Metrics within acceptable ranges.",
    "Operating efficiently."
]

os.makedirs("sample_reports", exist_ok=True)

# Styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
title_font = Font(bold=True, size=16, color="1E3A8A")
label_font = Font(bold=True, color="475569")
locked_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for p in plants:
    wb = Workbook()
    
    # 1. Instructions
    ws_inst = wb.active
    ws_inst.title = "Instructions"
    ws_inst.protection.sheet = True
    ws_inst.protection.password = 'dana123'
    
    ws_inst["A1"] = "DMEOP Enterprise Excel Intelligence Module"
    ws_inst["A1"].font = title_font
    instructions = [
        "Welcome to the DIBMS Automated Operations Reporting Module.", "",
        "HOW TO USE THIS TEMPLATE:",
        "1. Go to the 'Monthly Report' tab.",
        "2. Fill in the data for the requested reporting month.",
        "3. Only white cells can be edited. Gray cells are locked structural metadata.",
        "4. Revenue, Expenses, Production Units, and Safety Incidents must be numeric and non-negative.",
        "5. Attendance Rate and Quality Score must be between 0 and 100.",
        "6. Do not modify the hidden metadata in the Reference Data tab.",
        "7. Save the file and upload it via the DIBMS portal."
    ]
    for r_idx, text in enumerate(instructions, start=3):
        ws_inst[f"A{r_idx}"] = text
        ws_inst[f"A{r_idx}"].font = Font(color="333333") if not text.startswith("HOW TO") and text else Font(bold=True)
    ws_inst.column_dimensions['A'].width = 80
    
    # 2. Monthly Report
    ws_report = wb.create_sheet(title="Monthly Report")
    
    ws_report.protection.sheet = True
    ws_report.protection.password = 'dana123'
    
    ws_report["A1"] = "DANA INDIA OPERATIONS PORTAL"
    ws_report["A1"].font = title_font
    ws_report.merge_cells("A1:D1")
    
    metadata = [
        ("Plant ID", p["id"]),
        ("Plant Name", p["name"]),
        ("Manager Name", p["manager"]),
        ("Reporting Month", MONTH_NAMES[month]),
        ("Reporting Year", year),
    ]
    
    for r_idx, (label, val) in enumerate(metadata, start=3):
        ws_report[f"A{r_idx}"] = label
        ws_report[f"A{r_idx}"].font = label_font
        ws_report[f"A{r_idx}"].border = thin_border
        
        ws_report[f"B{r_idx}"] = val
        ws_report[f"B{r_idx}"].fill = locked_fill
        ws_report[f"B{r_idx}"].border = thin_border
        
    start_row = 10
    headers = ["Metric", "Value", "Remarks (Optional)"]
    for c_idx, header in enumerate(headers, start=1):
        cell = ws_report.cell(row=start_row, column=c_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        
    ws_report.freeze_panes = f"A{start_row+1}"
        
    metrics = [
        ("Revenue (INR)", "currency"),
        ("Expenses (INR)", "currency"),
        ("Production Units", "number"),
        ("Attendance Rate (%)", "percent"),
        ("Safety Incidents", "number"),
        ("Quality Score (%)", "percent")
    ]
    
    multiplier = 1.0 + (int(p["id"][-1]) * 0.1)
    revenue = round((8000000 * multiplier) + random.uniform(-500000, 500000), 2)
    expenses = round(revenue * random.uniform(0.65, 0.85), 2)
    prod_units = int((15000 * multiplier) + random.randint(-1000, 1000))
    attendance = round(random.uniform(90.0, 99.8), 2)
    quality = round(random.uniform(94.0, 99.8), 2)
    
    is_unsafe = random.random() > 0.8
    safety = random.randint(1, 2) if is_unsafe else 0
    
    vals = [revenue, expenses, prod_units, attendance, safety, quality]
    
    for r_idx, (metric_info, val) in enumerate(zip(metrics, vals), start=start_row + 1):
        metric_name, m_type = metric_info
        
        cell_label = ws_report.cell(row=r_idx, column=1)
        cell_label.value = metric_name
        cell_label.fill = locked_fill
        cell_label.border = thin_border
        cell_label.font = Font(bold=True)
        
        cell_val = ws_report.cell(row=r_idx, column=2)
        cell_val.value = val
        cell_val.border = thin_border
        cell_val.protection = openpyxl.styles.Protection(locked=False)
        
        if m_type == "currency":
            cell_val.number_format = '₹ #,##0.00'
        elif m_type == "percent":
            cell_val.number_format = '0.00'
            
        cell_rem = ws_report.cell(row=r_idx, column=3)
        cell_rem.value = random.choice(remarks_pool)
        cell_rem.border = thin_border
        cell_rem.protection = openpyxl.styles.Protection(locked=False)
        
    ws_report.column_dimensions['A'].width = 25
    ws_report.column_dimensions['B'].width = 20
    ws_report.column_dimensions['C'].width = 40
    
    # 3. Reference Data
    ws_ref = wb.create_sheet(title="Reference Data")
    ws_ref.sheet_state = 'hidden'
    
    ws_ref["A1"] = "Template Version"
    ws_ref["B1"] = template_version
    ws_ref["A2"] = "Plant ID"
    ws_ref["B2"] = p["id"]
    ws_ref["A3"] = "Month"
    ws_ref["B3"] = month
    ws_ref["A4"] = "Year"
    ws_ref["B4"] = year
    
    data_str = f"{template_version}|{p['id']}|{month}|{year}"
    checksum = hashlib.sha256(data_str.encode()).hexdigest()
    ws_ref["A5"] = "Checksum"
    ws_ref["B5"] = checksum
    
    filename = f"sample_reports/DIBMS_Template_{p['id']}_{year}_{month:02d}.xlsx"
    wb.save(filename)
    
print("Generated 8 sample reports in the sample_reports directory.")
